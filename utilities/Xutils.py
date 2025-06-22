import openpyxl


def get_row_count(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_read_data(file_name, sheet_name, row, col):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row, column=col).value


def get_write_data(file_name, sheet_name, row, col,data):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    sheet.cell(row=row, column=col).value = data
    workbook.save(file_name)
